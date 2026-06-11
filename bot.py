import os, re, json, logging, uuid
from datetime import datetime, timedelta
from collections import defaultdict
import anthropic
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import io

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "YOUR_TOKEN_HERE")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "YOUR_ANTHROPIC_KEY")
ADMIN_ID = 1885883892

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

expenses = []
incomes = []
balance_snapshots = []  # {date, cash, bank, chat_id} — остатки которые скидывает админ
pending = {}
ai_history = []  # история диалога с AI-ассистентом
chat_last_date = {}  # {chat_id: "ДД.ММ.ГГГГ"} — последняя известная дата в каждом чате

settings = {
    "categories": [
        "мясо","напитки","кола","сырьё","оплата за газ","рис","ФОТ",
        "электроэнергия","налог","доход","соленые тилла","телефония",
        "хлеб","фрукты","благоустройство","эхсон","масло","рыба",
        "долг","савзи","мфй","посуда","налог ежемесячный",
        "транспорт","такси","дебит кредит","70%","30%",
        "Азиз","Султонмурод","прочее"
    ],
    "custom_rules": {
        "азиз":"Азиз","султон мурод":"Султонмурод",
        "султонмурод":"Султонмурод","шохрух":"70%"
    }
}

KNOWN_FOT = [
    "комолиддин","аброр","фахридин","ойбек","шавкат","жохонгир","тохир",
    "сагдина","мохи","мохларой","самандар","зиевуддин","сунат","даврон","хуршид",
    "зухрат","муслима","бозор","тогара","мойка","уборка","тозалаш","иш хаки"
]

EMOJI = {
    "мясо":"🥩","напитки":"🥤","кола":"🥤","сырьё":"🛒","оплата за газ":"🔥",
    "рис":"🍚","ФОТ":"👷","электроэнергия":"⚡","налог":"📋","доход":"💰",
    "телефония":"📱","хлеб":"🍞","фрукты":"🍎","благоустройство":"🌿",
    "масло":"🫙","рыба":"🐟","савзи":"🥕","транспорт":"🚗","такси":"🚕",
    "дебит кредит":"🏦","70%":"💼","30%":"💼","Азиз":"👤","Султонмурод":"👤",
    "прочее":"📌","эхсон":"📌","мфй":"📌","посуда":"🍽️"
}

# ─── КЛАВИАТУРЫ ───────────────────────────────────────────────────────────────
def main_kb(uid=None):
    kb = [
        [KeyboardButton("📊 Отчёт"), KeyboardButton("💸 Движение средств")],
        [KeyboardButton("💰 Баланс"), KeyboardButton("📋 Последние записи")],
        [KeyboardButton("🔍 Детализация"), KeyboardButton("📥 Excel")],
        [KeyboardButton("ℹ️ Помощь")]
    ]
    if uid == ADMIN_ID:
        kb.insert(2, [KeyboardButton("📈 Отчёт по приходам")])
        kb.append([KeyboardButton("⚙️ Настройки")])
    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

def admin_kb():
    kb = [
        [KeyboardButton("➕ Добавить категорию"), KeyboardButton("🔧 Добавить правило")],
        [KeyboardButton("📋 Список правил"), KeyboardButton("🔙 Главное меню")]
    ]
    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

def period_kb(prefix="rep"):
    kb = [
        [InlineKeyboardButton("Сегодня", callback_data=f"{prefix}:today"),
         InlineKeyboardButton("Вчера", callback_data=f"{prefix}:yesterday")],
        [InlineKeyboardButton("7 дней", callback_data=f"{prefix}:week"),
         InlineKeyboardButton("30 дней", callback_data=f"{prefix}:month30")],
        [InlineKeyboardButton("Текущий месяц", callback_data=f"{prefix}:curmonth"),
         InlineKeyboardButton("Прошлый месяц", callback_data=f"{prefix}:prevmonth")],
        [InlineKeyboardButton("📅 Выбрать даты вручную", callback_data=f"{prefix}:calendar")]
    ]
    return InlineKeyboardMarkup(kb)

def cat_detail_kb(chat_id):
    cats = list({e['category'] for e in expenses if e['chat_id'] == chat_id})
    kb = []; row = []
    for cat in sorted(cats):
        row.append(InlineKeyboardButton(cat, callback_data=f"det:{cat}"))
        if len(row) == 3: kb.append(row); row = []
    if row: kb.append(row)
    return InlineKeyboardMarkup(kb)

def cat_assign_kb(pid):
    kb = []; row = []
    for cat in settings["categories"]:
        row.append(InlineKeyboardButton(cat, callback_data=f"asgn:{pid}:{cat}"))
        if len(row) == 3: kb.append(row); row = []
    if row: kb.append(row)
    kb.append([InlineKeyboardButton("🗑 Пропустить", callback_data=f"skip:{pid}")])
    return InlineKeyboardMarkup(kb)

# ─── БАЛАНС ───────────────────────────────────────────────────────────────────
def get_balance(chat_id):
    cash_in = sum(i['amount'] for i in incomes if i['chat_id']==chat_id and i['type']=='cash')
    bank_in = sum(i['amount'] for i in incomes if i['chat_id']==chat_id and i['type']=='bank')
    cash_out = sum(e['amount'] for e in expenses if e['chat_id']==chat_id and e.get('payment_type','cash')=='cash')
    bank_out = sum(e['amount'] for e in expenses if e['chat_id']==chat_id and e.get('payment_type','bank')=='bank')
    return {
        'cash_in':cash_in,'bank_in':bank_in,
        'cash_out':cash_out,'bank_out':bank_out,
        'cash_balance':cash_in-cash_out,
        'bank_balance':bank_in-bank_out,
        'total':cash_in+bank_in-cash_out-bank_out
    }

def check_balance_discrepancy(chat_id, snapshot_cash, snapshot_bank):
    """Проверить сходится ли остаток с расчётным"""
    b = get_balance(chat_id)
    cash_diff = snapshot_cash - b['cash_balance']
    bank_diff = snapshot_bank - b['bank_balance']
    issues = []
    if abs(cash_diff) > 1000:
        sign = "не хватает" if cash_diff < 0 else "лишнее"
        issues.append(f"💵 Наличка: {sign} *{abs(cash_diff):,.0f} сум*\n  Расчётный остаток: {b['cash_balance']:,.0f}\n  Фактический: {snapshot_cash:,.0f}")
    if abs(bank_diff) > 1000:
        sign = "не хватает" if bank_diff < 0 else "лишнее"
        issues.append(f"🏦 Банк: {sign} *{abs(bank_diff):,.0f} сум*\n  Расчётный: {b['bank_balance']:,.0f}\n  Фактический: {snapshot_bank:,.0f}")
    return issues

# ─── ДВИЖЕНИЕ СРЕДСТВ ─────────────────────────────────────────────────────────
def make_cashflow(chat_id, d_from, d_to, title):
    exps = [e for e in expenses if e['chat_id']==chat_id]
    incs = [i for i in incomes if i['chat_id']==chat_id]

    def in_range(item):
        try: return d_from <= datetime.strptime(item['date'], "%d.%m.%Y") <= d_to
        except: return False

    period_exp = [e for e in exps if in_range(e)]
    period_inc = [i for i in incs if in_range(i)]

    total_in = sum(i['amount'] for i in period_inc)
    total_out = sum(e['amount'] for e in period_exp)
    net = total_in - total_out

    cash_in = sum(i['amount'] for i in period_inc if i['type']=='cash')
    bank_in = sum(i['amount'] for i in period_inc if i['type']=='bank')
    cash_out = sum(e['amount'] for e in period_exp if e.get('payment_type','cash')=='cash')
    bank_out = sum(e['amount'] for e in period_exp if e.get('payment_type','bank')=='bank')

    # Топ категорий расходов
    by_cat = defaultdict(float)
    for e in period_exp: by_cat[e['category']] += e['amount']
    top5 = sorted(by_cat.items(), key=lambda x: x[1], reverse=True)[:5]

    sign = "+" if net >= 0 else ""
    lines = [
        f"💸 *Движение средств*",
        f"📆 {title}",
        f"",
        f"╔══════════════════════╗",
        f"║  📈 ПРИХОД           ║",
        f"║  💵 Нал:  {cash_in:>12,.0f}  ║",
        f"║  🏦 Банк: {bank_in:>12,.0f}  ║",
        f"║  Итого: {total_in:>13,.0f}  ║",
        f"╠══════════════════════╣",
        f"║  📉 РАСХОД           ║",
        f"║  💵 Нал:  {cash_out:>12,.0f}  ║",
        f"║  🏦 Банк: {bank_out:>12,.0f}  ║",
        f"║  Итого: {total_out:>13,.0f}  ║",
        f"╠══════════════════════╣",
        f"║  {'🟢' if net>=0 else '🔴'} ИТОГ: {sign}{net:>13,.0f}  ║",
        f"╚══════════════════════╝",
    ]

    if top5:
        lines += ["", "📊 *Топ расходов:*"]
        for cat, amt in top5:
            pct = amt/total_out*100 if total_out else 0
            bar = "█" * int(pct/10) + "░" * (10-int(pct/10))
            lines.append(f"{EMOJI.get(cat,'📌')} {cat[:12]:<12} {amt:>10,.0f}")

    return "\n".join(lines)

# ─── ОТЧЁТ ────────────────────────────────────────────────────────────────────
def make_report(title, chat_id, d_from, d_to):
    filtered = []
    for e in expenses:
        if e['chat_id'] != chat_id: continue
        try:
            if d_from <= datetime.strptime(e['date'], "%d.%m.%Y") <= d_to:
                filtered.append(e)
        except: pass
    if not filtered: return "📭 Нет данных за период"
    by_cat = defaultdict(float)
    for e in filtered: by_cat[e['category']] += e['amount']
    total = sum(by_cat.values())
    lines = [f"📊 *{title}*", f"📆 {d_from.strftime('%d.%m')} — {d_to.strftime('%d.%m.%Y')}\n"]
    for cat, amt in sorted(by_cat.items(), key=lambda x: x[1], reverse=True):
        lines.append(f"{EMOJI.get(cat,'📌')} *{cat}*: {amt:,.0f} сум")
    lines += ["","━━━━━━━━━━━━━━━",f"💰 *ИТОГО: {total:,.0f} сум*",f"📝 Записей: {len(filtered)}"]
    return "\n".join(lines)

def get_period(key):
    now = datetime.now()
    mn = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
          7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
    if key=="today": return now.replace(hour=0,minute=0,second=0),now.replace(hour=23,minute=59,second=59),"Сегодня"
    elif key=="yesterday":
        y=now-timedelta(days=1); return y.replace(hour=0,minute=0,second=0),y.replace(hour=23,minute=59,second=59),"Вчера"
    elif key=="week": return (now-timedelta(days=7)).replace(hour=0,minute=0,second=0),now.replace(hour=23,minute=59,second=59),"7 дней"
    elif key=="month30": return (now-timedelta(days=30)).replace(hour=0,minute=0,second=0),now.replace(hour=23,minute=59,second=59),"30 дней"
    elif key=="curmonth": return now.replace(day=1,hour=0,minute=0,second=0),now.replace(hour=23,minute=59,second=59),mn[now.month]
    elif key=="prevmonth":
        f=now.replace(day=1)-timedelta(days=1)
        return f.replace(day=1,hour=0,minute=0,second=0),f.replace(hour=23,minute=59,second=59),f"Прошлый {mn[f.month]}"
    return None,None,None

# ─── ИЗВЛЕЧЕНИЕ ДАТЫ ИЗ ТЕКСТА ───────────────────────────────────────────────
def extract_date_from_text(text):
    """Ищет дату в тексте — поддерживает разные форматы"""
    patterns = [
        r'\b(\d{2})\.(\d{2})\.(\d{4})\b',           # 19.05.2026
        r'\b(\d{1,2})\.(\d{2})\.(\d{4})\b',          # 9.05.2026
        r'\b(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})\b', # 19-05-2026
        r'\b(\d{1,2})\s+(январ|феврал|март|апрел|май|июн|июл|август|сентябр|октябр|ноябр|декабр)[а-я]*\s*\.?\s*(\d{4})\b',
        r'\b(\d{1,2})\s+(январ|феврал|март|апрел|май|июн|июл|август|сентябр|октябр|ноябр|декабр)[а-я]*\b',
        # Узбекские месяцы
        r'\b(\d{1,2})\s*(yanvar|fevral|mart|aprel|may|iyun|iyul|avgust|sentabr|oktabr|noyabr|dekabr)[a-z]*\b',
    ]
    month_map = {
        'январ':1,'феврал':2,'март':3,'апрел':4,'май':5,'июн':6,
        'июл':7,'август':8,'сентябр':9,'октябр':10,'ноябр':11,'декабр':12,
        'yanvar':1,'fevral':2,'mart':3,'aprel':4,'may':5,'iyun':6,
        'iyul':7,'avgust':8,'sentabr':9,'oktabr':10,'noyabr':11,'dekabr':12,
    }
    text_lower = text.lower()
    now = datetime.now()

    # Числовые форматы
    m = re.search(r'\b(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})\b', text)
    if m:
        try:
            d = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            return d.strftime("%d.%m.%Y")
        except: pass

    # Формат ДД.ММ без года
    m = re.search(r'\b(\d{1,2})\.(\d{1,2})\b', text)
    if m:
        try:
            d = datetime(now.year, int(m.group(2)), int(m.group(1)))
            return d.strftime("%d.%m.%Y")
        except: pass

    # Словесные месяцы
    for rus_month, num in month_map.items():
        pattern = rf'\b(\d{{1,2}})\s+{rus_month}'
        m = re.search(pattern, text_lower)
        if m:
            try:
                year = now.year
                year_m = re.search(rf'{rus_month}[а-яa-z]*\s*(\d{{4}})', text_lower)
                if year_m: year = int(year_m.group(1))
                d = datetime(year, num, int(m.group(1)))
                return d.strftime("%d.%m.%Y")
            except: pass
    return None

# ─── AI ПАРСИНГ РАСХОДОВ ──────────────────────────────────────────────────────
def parse_with_ai(text, fallback_date=None):
    try:
        cats = ", ".join(settings["categories"])
        rules = "\n".join([f"- '{k}' → {v}" for k,v in settings["custom_rules"].items()])
        today = datetime.now().strftime("%d.%m.%Y")
        default_date = fallback_date or today
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1500,
            messages=[{"role":"user","content":f"""Ты помощник учёта расходов ресторана. Русский/узбекский.
Сегодня: {today}
ДАТА ПО УМОЛЧАНИЮ: {default_date}
Правила: {rules}
ФОТ имена: {', '.join(KNOWN_FOT)}

Верни ТОЛЬКО JSON массив:
[{{"type":"expense"/"income","amount":число,"category":"кат или null","description":"описание","date":"ДД.ММ.ГГГГ","payment_type":"cash"/"bank","uncertain":bool,"found":true}}]

Категории: {cats}
- Имена из ФОТ → ФОТ | Азиз → Азиз | Султон Мурод → Султонмурод | Шохрух → 70%
- Такси → такси | Транспорт/тахи → транспорт | Газ → оплата за газ
- Гўшт/мясо → мясо | Сут → сырьё | Нон → хлеб | Кола → кола
- Телефон/Алиса/Сим расход → телефония | Шётка/щётка → сырьё
- Лагмон хамир, хамир → сырьё | Кабель вай фай → телефония
- Мева бозор → фрукты | Кандиянер/кондиционер → прочее
- Пичок/нож → сырьё | Пепси/Колага/Кумир → напитки
- "приход"/"тушум"/"касса" + сумма → type:income, payment_type:cash
- "банк приход"/"банк келди"/"Гушга" + сумма → type:income
- "банк" в расходе → payment_type:bank
- ВАЖНО: если в тексте НЕТ даты — используй ДАТУ ПО УМОЛЧАНИЮ: {default_date}
- Если дата ЕСТЬ в тексте — используй её для ВСЕХ строк этого сообщения
- Точки = тысячи: 50.000=50000 | Формула 57*3500 → результат
- Прочерк без суммы → пропустить | "ужн"/"кейин"/"вибга ужн" → пропустить
- Если непонятно → uncertain:true

Сообщение: {text}"""}]
        )
        raw = re.sub(r'```json|```','',resp.content[0].text.strip()).strip()
        result = json.loads(raw)
        return result if isinstance(result,list) else []
    except Exception as e:
        logger.error(f"AI parse: {e}"); return []

# ─── AI АССИСТЕНТ (личка с админом) ──────────────────────────────────────────
async def ai_assistant(update: Update, context, text: str):
    """Свободный диалог с AI-ассистентом в личке"""
    chat_id = update.effective_chat.id
    b = get_balance(chat_id)
    recent_exp = sorted(expenses, key=lambda x: x['date'], reverse=True)[:10]
    recent_inc = sorted(incomes, key=lambda x: x['date'], reverse=True)[:5]

    exp_summary = "\n".join([f"- {e['date']}: {e['category']} {e['amount']:,.0f} ({e['description'][:30]})" for e in recent_exp])
    inc_summary = "\n".join([f"- {i['date']}: {'🏦' if i['type']=='bank' else '💵'} +{i['amount']:,.0f} ({i['description'][:30]})" for i in recent_inc])
    rules_str = json.dumps(settings["custom_rules"], ensure_ascii=False)
    cats_str = ", ".join(settings["categories"])
    today = datetime.now().strftime("%d.%m.%Y")
    history = ai_history[-8:] if len(ai_history) > 8 else ai_history

    messages = history + [{"role":"user","content":f"""Сообщение от Жасура: {text}

Текущие данные:
Баланс наличка: {b['cash_balance']:,.0f} | Баланс банк: {b['bank_balance']:,.0f}
Последние расходы:
{exp_summary}
Последние приходы:
{inc_summary}
Категории: {cats_str}
Правила: {rules_str}
Сегодня: {today}"""}]

    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1000,
            system=f"""Ты финансовый ассистент ресторана Суром (Узбекистан). Помогаешь Жасуру управлять финансами.
Общайся на русском языке. Будь краток и конкретен.

ТВОИ ВОЗМОЖНОСТИ:
1. Записать расход/приход → верни в ответе: ACTION_EXP:{{...}} или ACTION_INC:{{...}}
2. Изменить категорию → верни: ACTION_RULE:{{\"keyword\":\"...\",\"category\":\"...\"}}
3. Если что-то непонятно → задай ОДИН уточняющий вопрос
4. Анализировать финансы и давать советы

ФОРМАТЫ ACTION:
ACTION_EXP:{{"amount":сумма,"category":"кат","description":"описание","date":"ДД.ММ.ГГГГ","payment_type":"cash"}}
ACTION_INC:{{"amount":сумма,"type":"cash"/"bank","description":"описание","date":"ДД.ММ.ГГГГ"}}
ACTION_RULE:{{"keyword":"слово","category":"категория"}}

ПРИМЕРЫ ПОНИМАНИЯ:
- "запиши расход мясо 500000 за 24 мая" → ACTION_EXP с category:мясо, date:24.05.{today[-4:]}
- "гушга 10800000 это приход в банк" → ACTION_INC с type:bank
- "азиз 140000 это не приход, это ФОТ" → ACTION_EXP с category:ФОТ
- "добавь правило: Мадина = 70%" → ACTION_RULE
- "какой баланс" → текстовый ответ с анализом
- "что за расход Бозор" → спроси уточнение если непонятно

ПРАВИЛО УТОЧНЕНИЙ: если не понимаешь сумму, дату или категорию — задай ОДИН конкретный вопрос.
Не задавай несколько вопросов сразу.""",
            messages=messages
        )
        answer = resp.content[0].text

        # Сохраняем историю
        ai_history.append({"role":"user","content":text})
        ai_history.append({"role":"assistant","content":answer})

        # Обрабатываем ACTION_EXP
        exp_match = re.search(r'ACTION_EXP:(\{[^}]+\})', answer)
        if exp_match:
            try:
                action = json.loads(exp_match.group(1))
                action['chat_id'] = ADMIN_ID
                action.setdefault('user', 'Жасур')
                expenses.append(action)
                answer = re.sub(r'ACTION_EXP:\{[^}]+\}', '', answer).strip()
                cat = action.get('category','')
                answer += f"\n\n✅ *Расход записан!*\n{EMOJI.get(cat,'📌')} {cat}: {float(action.get('amount',0)):,.0f} сум"
            except: pass

        # Обрабатываем ACTION_INC
        inc_match = re.search(r'ACTION_INC:(\{[^}]+\})', answer)
        if inc_match:
            try:
                action = json.loads(inc_match.group(1))
                action['chat_id'] = ADMIN_ID
                incomes.append(action)
                answer = re.sub(r'ACTION_INC:\{[^}]+\}', '', answer).strip()
                tp = "🏦 Банк" if action.get('type')=='bank' else "💵 Наличка"
                answer += f"\n\n✅ *Приход записан!*\n{tp}: {float(action.get('amount',0)):,.0f} сум"
            except: pass

        # Обрабатываем ACTION_RULE
        rule_match = re.search(r'ACTION_RULE:(\{[^}]+\})', answer)
        if rule_match:
            try:
                rule = json.loads(rule_match.group(1))
                settings["custom_rules"][rule['keyword'].lower()] = rule['category']
                answer = re.sub(r'ACTION_RULE:\{[^}]+\}', '', answer).strip()
                answer += f"\n\n✅ *Правило добавлено:* `{rule['keyword']}` → {rule['category']}"
            except: pass

        await update.message.reply_text(answer.strip(), parse_mode="Markdown")

    except Exception as e:
        logger.error(f"AI assistant: {e}")
        await update.message.reply_text("❌ Ошибка. Попробуйте ещё раз.")

# ─── EXCEL ────────────────────────────────────────────────────────────────────
def generate_excel(ce, ci):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        thin = Side(style='thin',color="CCCCCC"); border = Border(left=thin,right=thin,top=thin,bottom=thin)

        def header_row(ws, headers, color="C00000"):
            hf = PatternFill("solid",start_color=color,end_color=color)
            for c,h in enumerate(headers,1):
                cell = ws.cell(row=1,column=c,value=h)
                cell.fill=hf; cell.font=Font(bold=True,color="FFFFFF",name="Arial",size=10)
                cell.alignment=Alignment(horizontal="center"); cell.border=border

        ws = wb.active; ws.title="Расходы"
        header_row(ws,["Дата","Категория","Сумма","Тип","Описание","Кто"])
        for r,e in enumerate(sorted(ce,key=lambda x:x['date']),2):
            af = PatternFill("solid",start_color="FFF2CC" if r%2==0 else "FFFFFF",end_color="FFF2CC" if r%2==0 else "FFFFFF")
            for c,v in enumerate([e['date'],e['category'],e['amount'],e.get('payment_type','cash'),e['description'],e.get('user','')],1):
                cell=ws.cell(row=r,column=c,value=v); cell.font=Font(name="Arial",size=9); cell.fill=af; cell.border=border
                if c==3: cell.number_format='#,##0'

        ws2=wb.create_sheet("Приходы")
        header_row(ws2,["Дата","Тип","Сумма","Описание"],"375623")
        for r,i in enumerate(sorted(ci,key=lambda x:x['date']),2):
            for c,v in enumerate([i['date'],i['type'],i['amount'],i['description']],1):
                cell=ws2.cell(row=r,column=c,value=v)
                if c==3: cell.number_format='#,##0'

        ws3=wb.create_sheet("Итоги")
        header_row(ws3,["Категория","Сумма"],"1F4E79")
        by_cat=defaultdict(float)
        for e in ce: by_cat[e['category']]+=e['amount']
        total=sum(by_cat.values())
        for r,(cat,amt) in enumerate(sorted(by_cat.items(),key=lambda x:x[1],reverse=True),2):
            ws3.cell(row=r,column=1,value=cat); c=ws3.cell(row=r,column=2,value=amt); c.number_format='#,##0'
        ws3.cell(row=len(by_cat)+3,column=1,value="ИТОГО").font=Font(bold=True)
        tc=ws3.cell(row=len(by_cat)+3,column=2,value=total); tc.font=Font(bold=True); tc.number_format='#,##0'

        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
    except Exception as e:
        logger.error(f"Excel: {e}"); return None

# ─── ХЕНДЛЕРЫ ─────────────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    is_private = update.effective_chat.type == "private"
    if is_private and uid == ADMIN_ID:
        await update.message.reply_text(
            "☪️ *Ассаламу алейкум, Жасур!*\n\n"
            "Я твой личный финансовый ассистент 🤖\n\n"
            "В личном чате ты можешь писать мне *свободным текстом*:\n\n"
            "💬 *Примеры:*\n"
            "• `Запиши расход Мясо 500 000 за 24 мая`\n"
            "• `Покажи расходы за май`\n"
            "• `Какой у нас баланс?`\n"
            "• `Азиз теперь категория ФОТ`\n"
            "• `Остаток наличка 2 500 000, банк 8 000 000`\n\n"
            "Используй кнопки или пиши свободно 👇",
            parse_mode="Markdown", reply_markup=main_kb(uid))
    else:
        await update.message.reply_text(
            "☪️ *Ассаламу алейкум!*\n\n"
            "Я — *профессиональный бот-финансист* 🤖💼\n\n"
            "Создан Жасуром. Просто отправляйте отчёты!\n\n"
            "📝 Пример:\n"
            "```\n24.05.2026\nГўшт 50.000\nАброр 350.000\nТакси 30.000\n```",
            parse_mode="Markdown", reply_markup=main_kb(uid))

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    uid = update.effective_user.id
    chat_id = update.effective_chat.id
    is_private = update.effective_chat.type == "private"
    if not text or update.message.from_user.is_bot: return

    # В группе кнопки только для админа
    is_group = update.effective_chat.type in ["group","supergroup"]
    menu_buttons = ["📊 Отчёт","💸 Движение средств","💰 Баланс","📋 Последние записи","🔍 Детализация","📥 Excel","ℹ️ Помощь","⚙️ Настройки","➕ Добавить категорию","🔧 Добавить правило","📋 Список правил","🔙 Главное меню"]
    if is_group and uid != ADMIN_ID and text in menu_buttons: return

    # ─── Личный чат с админом — AI ассистент ───
    if is_private and uid == ADMIN_ID:
        # Проверяем на остаток (сверка баланса)
        snapshot_match = re.search(r'остаток[:\s]+нал[а-я]*[\s:]+([0-9\s.,]+).*банк[\s:]+([0-9\s.,]+)', text.lower())
        if snapshot_match:
            try:
                cash = float(re.sub(r'[^\d]','',snapshot_match.group(1)))
                bank = float(re.sub(r'[^\d]','',snapshot_match.group(2)))
                issues = check_balance_discrepancy(chat_id, cash, bank)
                balance_snapshots.append({'date':datetime.now().strftime("%d.%m.%Y"),'cash':cash,'bank':bank,'chat_id':chat_id})
                if issues:
                    msg = "⚠️ *Обнаружены расхождения!*\n\n" + "\n\n".join(issues)
                    msg += "\n\n💡 Возможно пропущены какие-то расходы или доходы."
                else:
                    b = get_balance(chat_id)
                    msg = f"✅ *Баланс сходится!*\n\n💵 Наличка: {cash:,.0f} сум\n🏦 Банк: {bank:,.0f} сум"
                await update.message.reply_text(msg, parse_mode="Markdown"); return
            except: pass

        # Обычные кнопки тоже работают в личке
        if text not in menu_buttons:
            await ai_assistant(update, context, text); return

    # ─── Кнопки меню ───
    if text == "📊 Отчёт":
        await update.message.reply_text("Выберите период:", reply_markup=period_kb()); return
    elif text == "💸 Движение средств":
        await update.message.reply_text("Выберите период:", reply_markup=period_kb("flow")); return
    elif text == "💰 Баланс":
        await show_balance(update, context); return
    elif text == "📋 Последние записи":
        await show_list(update, context); return
    elif text == "📈 Отчёт по приходам":
        if uid != ADMIN_ID: return
        await show_income_report(update, context); return
        kb = cat_detail_kb(chat_id)
        if kb.inline_keyboard: await update.message.reply_text("Выберите категорию:", reply_markup=kb)
        else: await update.message.reply_text("📭 Нет записей.")
        return
    elif text == "📥 Excel":
        await download_excel(update, context); return
    elif text == "ℹ️ Помощь":
        await show_help(update, context); return
    elif text == "🔙 Главное меню":
        await update.message.reply_text("Главное меню:", reply_markup=main_kb(uid)); return
    elif text == "⚙️ Настройки":
        if uid != ADMIN_ID: return
        await update.message.reply_text("⚙️ *Панель администратора*", parse_mode="Markdown", reply_markup=admin_kb()); return
    elif text == "➕ Добавить категорию":
        if uid != ADMIN_ID: return
        context.user_data['state'] = 'add_cat'
        await update.message.reply_text("Введите название новой категории:"); return
    elif text == "🔧 Добавить правило":
        if uid != ADMIN_ID: return
        context.user_data['state'] = 'add_rule'
        await update.message.reply_text("Введите ключевое слово:\n_(например: `Мадина`)_", parse_mode="Markdown"); return
    elif text == "📋 Список правил":
        if uid != ADMIN_ID: return
        cats = "\n".join([f"• {c}" for c in settings["categories"]])
        rules = "\n".join([f"• `{k}` → *{v}*" for k,v in settings["custom_rules"].items()]) or "_(пусто)_"
        await update.message.reply_text(f"📋 *Категории:*\n{cats}\n\n🔧 *Правила:*\n{rules}", parse_mode="Markdown", reply_markup=admin_kb()); return

    # ─── Диалог ───
    state = context.user_data.get('state')
    if state == 'add_cat' and uid == ADMIN_ID:
        nc = text.strip()
        if nc not in settings["categories"]: settings["categories"].append(nc)
        await update.message.reply_text(f"✅ *{nc}* добавлена!", parse_mode="Markdown", reply_markup=admin_kb())
        context.user_data['state'] = None; return
    elif state == 'add_rule' and uid == ADMIN_ID:
        context.user_data['rule_kw'] = text.strip().lower()
        kb = []; row = []
        for cat in settings["categories"]:
            row.append(InlineKeyboardButton(cat, callback_data=f"setrule:{cat}"))
            if len(row) == 3: kb.append(row); row = []
        if row: kb.append(row)
        await update.message.reply_text(f"Категория для *{text.strip()}*?", parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(kb)); return

    if text.startswith('/'): return

    # ─── Обновляем последнюю известную дату для чата ───
    found_date = extract_date_from_text(text)
    if found_date:
        chat_last_date[chat_id] = found_date

    # Дата по умолчанию = последняя известная дата чата
    fallback = chat_last_date.get(chat_id)

    # ─── Обработка расходов ───
    items = parse_with_ai(text, fallback_date=fallback)
    if not items: return

    saved_exp, saved_inc, uncertain_items = [], [], []
    for item in items:
        if not item.get('found') or not item.get('amount'): continue
        item_date = item.get('date', datetime.now().strftime("%d.%m.%Y"))
        try: datetime.strptime(item_date, "%d.%m.%Y")
        except: item_date = datetime.now().strftime("%d.%m.%Y")

        if item.get('type') == 'income':
            inc = {'date':item_date,'amount':float(item['amount']),'type':item.get('payment_type','cash'),'description':item.get('description','')[:100],'chat_id':chat_id}
            incomes.append(inc); saved_inc.append(inc)
        else:
            exp = {'date':item_date,'amount':float(item['amount']),'category':item.get('category') or 'прочее','description':item.get('description','')[:100],'user':update.message.from_user.first_name or "—",'chat_id':chat_id,'payment_type':item.get('payment_type','cash')}
            if item.get('uncertain') or not item.get('category'): uncertain_items.append(exp)
            else: expenses.append(exp); saved_exp.append(exp)

    lines = []
    for i in saved_inc:
        tp = "🏦 Банк" if i['type']=='bank' else "💵 Нал"
        lines.append(f"✅ *Приход {tp}*: {i['amount']:,.0f} сум ({i['date']})")
    if saved_exp:
        if len(saved_exp) == 1:
            e = saved_exp[0]; pt = "🏦" if e.get('payment_type')=='bank' else "💵"
            lines.append(f"✅ {EMOJI.get(e['category'],'📌')} *{e['category']}* {pt}: {e['amount']:,.0f} сум | {e['date']}")
        else:
            lines.append(f"✅ *Записано {len(saved_exp)} позиций:*")
            for e in saved_exp:
                lines.append(f"{EMOJI.get(e['category'],'📌')} {e['description']}: {e['amount']:,.0f} → *{e['category']}* ({e['date']})")
            lines.append(f"💰 *Итого: {sum(e['amount'] for e in saved_exp):,.0f} сум*")
    if lines:
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

    for exp in uncertain_items:
        pid = str(uuid.uuid4())[:8]; pending[pid] = exp
        try:
            await context.bot.send_message(chat_id=ADMIN_ID,
                text=f"❓ *Не понял категорию!*\n\n📝 *{exp['description']}*\n💵 {exp['amount']:,.0f} сум | 📆 {exp['date']}\n👤 {exp['user']}\n\nВыберите категорию:",
                parse_mode="Markdown", reply_markup=cat_assign_kb(pid))
        except Exception as e:
            exp['category'] = 'прочее'; expenses.append(exp)

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    d = q.data; chat_id = q.message.chat_id; uid = q.from_user.id

    # ─── Отчёт по периоду ───
    if d.startswith("rep:"):
        key = d.replace("rep:","")
        if key == "calendar":
            now = datetime.now()
            context.user_data['cal_mode'] = 'rep'
            context.user_data['cal_step'] = 'from'
            await q.edit_message_text("📅 Выберите *начальную* дату:", parse_mode="Markdown",
                reply_markup=calendar_kb(now.year, now.month, "from"))
            return
        df,dt,title = get_period(key)
        if df: await q.edit_message_text(make_report(title,chat_id,df,dt), parse_mode="Markdown")

    elif d.startswith("flow:"):
        key = d.replace("flow:","")
        if key == "calendar":
            now = datetime.now()
            context.user_data['cal_mode'] = 'flow'
            context.user_data['cal_step'] = 'from'
            await q.edit_message_text("📅 Выберите *начальную* дату:", parse_mode="Markdown",
                reply_markup=calendar_kb(now.year, now.month, "from"))
            return
        df,dt,title = get_period(key)
        if df: await q.edit_message_text(make_cashflow(chat_id,df,dt,title), parse_mode="Markdown")

    # ─── Календарь ───
    elif d.startswith("cal:"):
        parts = d.split(":")
        action = parts[1]

        if action == "ignore": return
        if action == "cancel":
            await q.edit_message_text("❌ Отменено.")
            context.user_data.pop('cal_mode', None)
            context.user_data.pop('cal_step', None)
            context.user_data.pop('cal_from', None)
            return

        if action in ("prev", "next"):
            year, month = int(parts[2]), int(parts[3])
            mode = parts[4]
            if action == "prev":
                month -= 1
                if month < 1: month = 12; year -= 1
            else:
                month += 1
                if month > 12: month = 1; year += 1
            label = "начальную" if mode == "from" else "конечную"
            await q.edit_message_text(f"📅 Выберите *{label}* дату:", parse_mode="Markdown",
                reply_markup=calendar_kb(year, month, mode))
            return

        if action == "day":
            year, month, day = int(parts[2]), int(parts[3]), int(parts[4])
            mode = parts[5]
            selected = datetime(year, month, day)
            selected_str = selected.strftime("%d.%m.%Y")

            if mode == "from":
                context.user_data['cal_from'] = selected
                context.user_data['cal_step'] = 'to'
                await q.edit_message_text(
                    f"✅ Начало: *{selected_str}*\n\n📅 Теперь выберите *конечную* дату:",
                    parse_mode="Markdown",
                    reply_markup=calendar_kb(year, month, "to"))
            else:
                # Конечная дата выбрана — строим отчёт
                d_from = context.user_data.get('cal_from')
                d_to = selected.replace(hour=23, minute=59, second=59)
                cal_mode = context.user_data.get('cal_mode', 'rep')

                if d_from and d_to >= d_from:
                    title = f"{d_from.strftime('%d.%m')} — {d_to.strftime('%d.%m.%Y')}"
                    if cal_mode == 'rep':
                        report = make_report(title, chat_id, d_from, d_to)
                    else:
                        report = make_cashflow(chat_id, d_from, d_to, title)
                    await q.edit_message_text(report, parse_mode="Markdown")
                else:
                    await q.edit_message_text("❌ Конечная дата должна быть позже начальной. Попробуйте снова.")

                context.user_data.pop('cal_mode', None)
                context.user_data.pop('cal_step', None)
                context.user_data.pop('cal_from', None)

    # ─── Детализация ───
    elif d.startswith("det:"):
        cat = d.replace("det:","")
        cat_exps = [e for e in expenses if e['chat_id']==chat_id and e['category']==cat]
        if not cat_exps: await q.edit_message_text(f"📭 Нет записей в *{cat}*", parse_mode="Markdown"); return
        total = sum(e['amount'] for e in cat_exps)
        lines = [f"🔍 *Детализация: {cat}*\n"]
        for e in sorted(cat_exps, key=lambda x:x['date'], reverse=True)[:20]:
            pt = "🏦" if e.get('payment_type')=='bank' else "💵"
            lines.append(f"📆 {e['date']} {pt} {e['amount']:,.0f} | {e['description'][:30]}")
        lines += ["","━━━━━━━━━━━━━━━",f"💰 *Итого: {total:,.0f} сум* ({len(cat_exps)} записей)"]
        await q.edit_message_text("\n".join(lines), parse_mode="Markdown")

    elif d.startswith("asgn:"):
        parts=d.split(":",2); pid=parts[1]; cat=parts[2]
        if pid in pending:
            exp=pending.pop(pid); exp['category']=cat; expenses.append(exp)
            kw = exp['description'].lower().split()[0] if exp['description'] else ''
            if kw and len(kw)>2: settings["custom_rules"][kw]=cat
            await q.edit_message_text(f"✅ {EMOJI.get(cat,'📌')} *{cat}*\n💵 {exp['amount']:,.0f} сум\n📌 Правило: `{kw}` → {cat}", parse_mode="Markdown")

    elif d.startswith("skip:"):
        pid=d.replace("skip:","")
        if pid in pending: pending.pop(pid)
        await q.edit_message_text("🗑 Пропущено.")

    elif d.startswith("setrule:"):
        cat=d.replace("setrule:",""); kw=context.user_data.get('rule_kw','')
        if kw: settings["custom_rules"][kw]=cat
        await q.edit_message_text(f"✅ `{kw}` → *{cat}*", parse_mode="Markdown")
        context.user_data['state']=None; context.user_data['rule_kw']=None

async def show_income_report(update, context):
    chat_id = update.effective_chat.id
    if not incomes and not expenses:
        await update.message.reply_text("📭 Нет данных."); return

    ci = [i for i in incomes if i['chat_id'] == chat_id]
    ce = [e for e in expenses if e['chat_id'] == chat_id]

    # Приходы по типу
    cash_in = sum(i['amount'] for i in ci if i['type'] == 'cash')
    bank_in = sum(i['amount'] for i in ci if i['type'] == 'bank')
    total_in = cash_in + bank_in

    # Расходы по типу
    cash_out = sum(e['amount'] for e in ce if e.get('payment_type','cash') == 'cash')
    bank_out = sum(e['amount'] for e in ce if e.get('payment_type','bank') == 'bank')
    total_out = cash_out + bank_out

    net = total_in - total_out

    # Последние приходы
    recent_inc = sorted(ci, key=lambda x: x['date'], reverse=True)[:10]
    inc_lines = "\n".join([
        f"📆 {i['date']} | {'🏦' if i['type']=='bank' else '💵'} {i['amount']:,.0f} | {i['description'][:30]}"
        for i in recent_inc
    ]) or "_(нет записей)_"

    sign = "+" if net >= 0 else ""
    color = "🟢" if net >= 0 else "🔴"

    text = (
        f"📈 *Отчёт по приходам*\n\n"
        f"*ПРИХОДЫ:*\n"
        f"💵 Наличка: {cash_in:,.0f} сум\n"
        f"🏦 Банк: {bank_in:,.0f} сум\n"
        f"📥 *Итого приход: {total_in:,.0f} сум*\n\n"
        f"*РАСХОДЫ:*\n"
        f"💵 Наличка: {cash_out:,.0f} сум\n"
        f"🏦 Банк: {bank_out:,.0f} сум\n"
        f"📤 *Итого расход: {total_out:,.0f} сум*\n\n"
        f"━━━━━━━━━━━━━━━\n"
        f"{color} *Чистый остаток: {sign}{net:,.0f} сум*\n\n"
        f"*Последние приходы:*\n{inc_lines}"
    )
    await update.message.reply_text(text, parse_mode="Markdown")


    b = get_balance(update.effective_chat.id)
    cash_sign = "🟢" if b['cash_balance']>=0 else "🔴"
    bank_sign = "🟢" if b['bank_balance']>=0 else "🔴"
    total_sign = "🟢" if b['total']>=0 else "🔴"
    await update.message.reply_text(
        f"💰 *Баланс*\n\n"
        f"💵 *Наличка:*\n"
        f"  📈 Приход: {b['cash_in']:,.0f} сум\n"
        f"  📉 Расход: {b['cash_out']:,.0f} сум\n"
        f"  {cash_sign} Остаток: *{b['cash_balance']:,.0f} сум*\n\n"
        f"🏦 *Банк:*\n"
        f"  📈 Приход: {b['bank_in']:,.0f} сум\n"
        f"  📉 Расход: {b['bank_out']:,.0f} сум\n"
        f"  {bank_sign} Остаток: *{b['bank_balance']:,.0f} сум*\n\n"
        f"━━━━━━━━━━━━━━━\n"
        f"{total_sign} *Общий остаток: {b['total']:,.0f} сум*",
        parse_mode="Markdown")

async def show_list(update, context):
    ce = [e for e in expenses if e['chat_id']==update.effective_chat.id]
    recent = sorted(ce, key=lambda x:x['date'], reverse=True)[:10]
    if not recent: await update.message.reply_text("📭 Нет записей."); return
    lines = ["📋 *Последние 10:*\n"]
    for e in recent:
        pt = "🏦" if e.get('payment_type')=='bank' else "💵"
        lines.append(f"{EMOJI.get(e['category'],'📌')}{pt} {e['date']} | *{e['category']}* | {e['amount']:,.0f}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def download_excel(update, context):
    chat_id = update.effective_chat.id
    ce = [e for e in expenses if e['chat_id']==chat_id]
    ci = [i for i in incomes if i['chat_id']==chat_id]
    if not ce and not ci: await update.message.reply_text("📭 Нет данных."); return
    await update.message.reply_text("⏳ Генерирую Excel...")
    buf = generate_excel(ce, ci)
    if buf:
        await update.message.reply_document(document=buf,
            filename=f"расходы_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
            caption=f"📊 {len(ce)} расходов, {len(ci)} приходов")
    else: await update.message.reply_text("❌ Ошибка.")

async def show_help(update, context):
    await update.message.reply_text(
        "📖 *Инструкция*\n\n"
        "*Расходы в группе:*\n```\n24.05.2026\nГўшт 50.000\nАброр 350.000\n```\n\n"
        "*Приход наличкой:* `Приход 5 000 000`\n"
        "*Приход в банк:* `Банк приход 10 000 000`\n"
        "*Расход с банка:* `Аренда банк 1 500 000`\n\n"
        "*Сверка остатка (в личке боту):*\n"
        "`Остаток наличка 2500000 банк 8000000`\n"
        "Бот проверит совпадает ли с расчётом!\n\n"
        "📊 *Отчёт* — по категориям\n"
        "💸 *Движение* — приход/расход/итог\n"
        "🔍 *Детализация* — по одной категории\n"
        "💬 *В личке* — пиши боту свободно как ассистенту",
        parse_mode="Markdown", reply_markup=main_kb(update.effective_user.id))

def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.TEXT, handle_message))
    logger.info("Bot started!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()

# ═══════════════════════════════════════════════════════════════════
# CALENDAR MODULE — мини-календарь для выбора периода
# ═══════════════════════════════════════════════════════════════════

def calendar_kb(year, month, mode="from"):
    """Генерирует inline-клавиатуру с календарём"""
    import calendar
    month_names = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
                   7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
    kb = []
    # Заголовок с навигацией
    kb.append([
        InlineKeyboardButton("◀️", callback_data=f"cal:prev:{year}:{month}:{mode}"),
        InlineKeyboardButton(f"{month_names[month]} {year}", callback_data="cal:ignore"),
        InlineKeyboardButton("▶️", callback_data=f"cal:next:{year}:{month}:{mode}")
    ])
    # Дни недели
    kb.append([InlineKeyboardButton(d, callback_data="cal:ignore") for d in ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]])
    # Дни месяца
    cal = calendar.monthcalendar(year, month)
    for week in cal:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="cal:ignore"))
            else:
                row.append(InlineKeyboardButton(str(day), callback_data=f"cal:day:{year}:{month}:{day}:{mode}"))
        kb.append(row)
    # Кнопка отмены
    kb.append([InlineKeyboardButton("❌ Отмена", callback_data="cal:cancel")])
    return InlineKeyboardMarkup(kb)

